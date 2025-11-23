# app_diario_comidas_v2_1.py
# -*- coding: utf-8 -*-

from __future__ import annotations
from datetime import datetime, date
from typing import List, Dict
import io

import pandas as pd
import streamlit as st

# =========================
# Configuración básica
# =========================
st.set_page_config(page_title="Diario de Comidas", layout="wide")
HORA_FMT = "%H:%M"

# =========================
# Estilos responsivos (móvil/tablet/desktop)
# =========================
st.markdown("""
<style>
.block-container {
    padding-top: 2.2rem;       /* antes .5rem → más espacio superior */
    padding-bottom: 2rem;
    max-width: 1200px;
}
h1 {
    line-height: 1.25;         /* evita que el emoji se recorte */
    overflow: visible !important;
}
.stProgress > div > div { height: 16px; border-radius: 12px; }
.stButton > button { padding: .65rem 1rem; border-radius: 12px; }
.stNumberInput input, .stTextInput input { border-radius: 10px; }

/* Móvil */
@media (max-width: 480px) {
  .block-container { padding-left: .6rem; padding-right: .6rem; padding-top: 1.5rem; }
  .stProgress > div > div { height: 22px; }
  .stButton > button { width: 100%; font-size: 1rem; }
  .stNumberInput label, .stTextInput label, .stSelectbox label { font-size: 0.95rem; }
  .stDataFrame { font-size: .92rem; }
  .st-emotion-cache-ocqkz7, .st-emotion-cache-1wmy9hl { display: block !important; }
}

/* Tablet */
@media (min-width: 481px) and (max-width: 1024px) {
  .stProgress > div > div { height: 20px; }
  .stButton > button { font-size: .95rem; }
}
</style>
""", unsafe_allow_html=True)

# =========================
# Logo superior derecho
# =========================
from pathlib import Path
import base64

logo_path = Path(__file__).parent / "Logohorizontal.png"

with open(logo_path, "rb") as f:
    logo_base64 = base64.b64encode(f.read()).decode()

st.markdown(
    f"""
    <style>
    [data-testid="stAppViewContainer"] {{
        position: relative;
    }}
    .logo-tribu {{
        position: absolute;
        top: 10px;
        right: 25px;
        width: 180px;
        z-index: 999;
    }}
    </style>
    <img src="data:image/png;base64,{logo_base64}" class="logo-tribu">
    """,
    unsafe_allow_html=True
)


# =========================
# BASE INTERNA (proporcionada)
# =========================
BASE_INTERNA: List[Dict] = [
  {
    "nombre": "Pechuga de pollo (cruda, sin piel)",
    "porcion_desc": "100g",
    "kcal": 120.0,
    "proteina_g": 22.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Carne roja magra (cruda)",
    "porcion_desc": "100g",
    "kcal": 170.0,
    "proteina_g": 27.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Carne roja (cruda)",
    "porcion_desc": "100g",
    "kcal": 280.0,
    "proteina_g": 23.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Pavo (pechuga cruda, sin piel)",
    "porcion_desc": "100g",
    "kcal": 110.0,
    "proteina_g": 21.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Hamburguesa (completa)",
    "porcion_desc": "1 unidad",
    "kcal": 615.0,
    "proteina_g": 32.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Arepa",
    "porcion_desc": "1 unidad",
    "kcal": 218.0,
    "proteina_g": 5.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Estofado de pollo",
    "porcion_desc": "1 plato",
    "kcal": 350.0,
    "proteina_g": 22.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Estofado de res",
    "porcion_desc": "1 plato",
    "kcal": 450.0,
    "proteina_g": 30.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Chuleta de cerdo magra (cruda)",
    "porcion_desc": "100g",
    "kcal": 143.0,
    "proteina_g": 21.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Salmón (crudo)",
    "porcion_desc": "100g",
    "kcal": 142.0,
    "proteina_g": 20.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Huevo Frtio",
    "porcion_desc": "1 unidad",
    "kcal": 110.0,
    "proteina_g": 7.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Paella",
    "porcion_desc": "1 plato",
    "kcal": 400.0,
    "proteina_g": 15.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Chicharron de cerdo",
    "porcion_desc": "1 plato",
    "kcal": 500.0,
    "proteina_g": 40.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Clara de huevo",
    "porcion_desc": "1 unidad",
    "kcal": 4.0,
    "proteina_g": 4.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Bowl de Ensalada",
    "porcion_desc": "1 unidad",
    "kcal": 170.0,
    "proteina_g": 4.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Atún fresco (crudo)",
    "porcion_desc": "100g",
    "kcal": 144.0,
    "proteina_g": 23.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Atún en lata (en aceite)",
    "porcion_desc": "1 unidad (120g)",
    "kcal": 250.0,
    "proteina_g": 26.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Atún en lata (en agua)",
    "porcion_desc": "1 unidad (120g)",
    "kcal": 120.0,
    "proteina_g": 28.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Merluza (cruda, pescado blanco)",
    "porcion_desc": "100g",
    "kcal": 85.0,
    "proteina_g": 18.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Bacalao (crudo, pescado blanco)",
    "porcion_desc": "100g",
    "kcal": 82.0,
    "proteina_g": 18.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Tilapia (cruda)",
    "porcion_desc": "100g",
    "kcal": 96.0,
    "proteina_g": 20.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Camarones (crudos)",
    "porcion_desc": "100g",
    "kcal": 99.0,
    "proteina_g": 24.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Huevo entero",
    "porcion_desc": "1 unidad",
    "kcal": 70.0,
    "proteina_g": 6.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Rapidita",
    "porcion_desc": "1 unidad",
    "kcal": 145.0,
    "proteina_g": 4.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Requesón (cottage 2%)",
    "porcion_desc": "100g",
    "kcal": 82.0,
    "proteina_g": 11.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Legumbres (cocidas, secas, crudas)",
    "porcion_desc": "100g",
    "kcal": 116.0,
    "proteina_g": 9.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Edamame (vaina verde, crudo)",
    "porcion_desc": "100g",
    "kcal": 122.0,
    "proteina_g": 11.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Tofu firme",
    "porcion_desc": "100g",
    "kcal": 76.0,
    "proteina_g": 8.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Batido Nutricional",
    "porcion_desc": "1 scoop",
    "kcal": 45.0,
    "proteina_g": 4.5,
    "hidr_ml": 0
  },
  {
    "nombre": "PDM",
    "porcion_desc": "1 scoop",
    "kcal": 55.0,
    "proteina_g": 9.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Olluquitos",
    "porcion_desc": "1 plato",
    "kcal": 195.0,
    "proteina_g": 3.0,
    "hidr_ml": 0
  },
  {
    "nombre": "H24 Rebuild Strength",
    "porcion_desc": "1 scoop",
    "kcal": 95.0,
    "proteina_g": 12.5,
    "hidr_ml": 0
  },
  {
    "nombre": "Crocante de Proteina",
    "porcion_desc": "1 barra",
    "kcal": 37.0,
    "proteina_g": 7.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Bebida Pro H24",
    "porcion_desc": "1 sobre",
    "kcal": 35.0,
    "proteina_g": 17.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Beverage Mix",
    "porcion_desc": "1 scoop",
    "kcal": 37.0,
    "proteina_g": 15.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Proteina Personalizada en Polvo (PPP)",
    "porcion_desc": "1 scoop (6g)",
    "kcal": 20.0,
    "proteina_g": 5.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Lentejas (cocidas)",
    "porcion_desc": "100g",
    "kcal": 116.0,
    "proteina_g": 9.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Garbanzos (cocidos)",
    "porcion_desc": "100g",
    "kcal": 164.0,
    "proteina_g": 9.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Arvejas (cocidas)",
    "porcion_desc": "100g",
    "kcal": 84.0,
    "proteina_g": 5.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Espinaca",
    "porcion_desc": "100g",
    "kcal": 23.0,
    "proteina_g": 3.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Brócoli",
    "porcion_desc": "100g",
    "kcal": 34.0,
    "proteina_g": 3.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Zanahoria",
    "porcion_desc": "100g",
    "kcal": 41.0,
    "proteina_g": 1.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Pepino",
    "porcion_desc": "100g",
    "kcal": 16.0,
    "proteina_g": 0.7,
    "hidr_ml": 0
  },
  {
    "nombre": "Cebolla",
    "porcion_desc": "100g",
    "kcal": 40.0,
    "proteina_g": 1.1,
    "hidr_ml": 0
  },
  {
    "nombre": "Pimiento",
    "porcion_desc": "100g",
    "kcal": 31.0,
    "proteina_g": 1.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Tomate",
    "porcion_desc": "100g",
    "kcal": 18.0,
    "proteina_g": 0.9,
    "hidr_ml": 0
  },
  {
    "nombre": "Plátano",
    "porcion_desc": "1 unidad",
    "kcal": 105.0,
    "proteina_g": 1.3,
    "hidr_ml": 0
  },
  {
    "nombre": "Manzana",
    "porcion_desc": "1 unidad",
    "kcal": 95.0,
    "proteina_g": 0.5,
    "hidr_ml": 0
  },
  {
    "nombre": "Pera",
    "porcion_desc": "1 unidad",
    "kcal": 101.0,
    "proteina_g": 0.6,
    "hidr_ml": 0
  },
  {
    "nombre": "Naranja",
    "porcion_desc": "1 unidad",
    "kcal": 62.0,
    "proteina_g": 1.2,
    "hidr_ml": 0
  },
  {
    "nombre": "Mandarina",
    "porcion_desc": "1 unidad",
    "kcal": 47.0,
    "proteina_g": 0.7,
    "hidr_ml": 0
  },
  {
    "nombre": "Piña",
    "porcion_desc": "1 taza (165g)",
    "kcal": 82.0,
    "proteina_g": 0.9,
    "hidr_ml": 0
  },
  {
    "nombre": "Papaya",
    "porcion_desc": "1 taza (145g)",
    "kcal": 62.0,
    "proteina_g": 0.7,
    "hidr_ml": 0
  },
  {
    "nombre": "Mango",
    "porcion_desc": "1 taza (165g)",
    "kcal": 99.0,
    "proteina_g": 1.4,
    "hidr_ml": 0
  },
  {
    "nombre": "Uvas",
    "porcion_desc": "1 puñado (80g)",
    "kcal": 55.0,
    "proteina_g": 0.6,
    "hidr_ml": 0
  },
  {
    "nombre": "Fresas",
    "porcion_desc": "1 puñado (100g)",
    "kcal": 33.0,
    "proteina_g": 0.7,
    "hidr_ml": 0
  },
  {
    "nombre": "Kiwi",
    "porcion_desc": "1 unidad",
    "kcal": 42.0,
    "proteina_g": 0.8,
    "hidr_ml": 0
  },
  {
    "nombre": "Almendras",
    "porcion_desc": "1 puñado (28g)",
    "kcal": 170.0,
    "proteina_g": 6.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Maní (cacahuate)",
    "porcion_desc": "1 puñado (28g)",
    "kcal": 165.0,
    "proteina_g": 7.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Nueces",
    "porcion_desc": "1 puñado (28g)",
    "kcal": 185.0,
    "proteina_g": 4.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Arroz Blanco (Cocido)",
    "porcion_desc": "1 taza (150g)",
    "kcal": 200.0,
    "proteina_g": 4.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Arroz Integral (Cocido)",
    "porcion_desc": "1 taza (150g)",
    "kcal": 215.0,
    "proteina_g": 5.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Pasta Blanca (Cocida)",
    "porcion_desc": "1 taza (150g)",
    "kcal": 220.0,
    "proteina_g": 7.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Papa (Cocida)",
    "porcion_desc": "1 unidad (150g)",
    "kcal": 130.0,
    "proteina_g": 3.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Camote (Cocido)",
    "porcion_desc": "1 unidad (150g)",
    "kcal": 135.0,
    "proteina_g": 2.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Quinua (Cocida)",
    "porcion_desc": "1 taza (185g)",
    "kcal": 222.0,
    "proteina_g": 8.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Avena (Cocida)",
    "porcion_desc": "1 taza (234g)",
    "kcal": 154.0,
    "proteina_g": 6.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Pasta Integral (Cocida)",
    "porcion_desc": "1 taza (150g)",
    "kcal": 210.0,
    "proteina_g": 8.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Aceituna verde deshuesada",
    "porcion_desc": "1 unidad",
    "kcal": 5,
    "proteina_g": 0.05,
    "hidr_ml": 0
  },
  {
    "nombre": "Herbal Aloe Concentrado",
    "porcion_desc": "3 tapas",
    "kcal": 1.8,
    "proteina_g": 0.0,
    "hidr_ml": 0
  },
  {
    "nombre": "N-R-G",
    "porcion_desc": "1/2 cucharadita",
    "kcal": 3.4,
    "proteina_g": 0.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Té Concentrado de Hierbas",
    "porcion_desc": "1/2 cucharadita",
    "kcal": 6.0,
    "proteina_g": 0.0,
    "hidr_ml": 0
  },
  {
    "nombre": "CR7 Drive",
    "porcion_desc": "1 scoop",
    "kcal": 50.0,
    "proteina_g": 0.0,
    "hidr_ml": 0
  },
  {
    "nombre": "RLX",
    "porcion_desc": "1 sobre",
    "kcal": 5.0,
    "proteina_g": 0.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Collagen Beauty Drink",
    "porcion_desc": "1 scoop",
    "kcal": 25.0,
    "proteina_g": 2.5,
    "hidr_ml": 0
  },
  {
    "nombre": "Niteworks",
    "porcion_desc": "1 scoop",
    "kcal": 25.0,
    "proteina_g": 0.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Immune Essentials",
    "porcion_desc": "1 scoop",
    "kcal": 25.0,
    "proteina_g": 0.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Complex Multivitamínico",
    "porcion_desc": "1 cápsula",
    "kcal": 5.0,
    "proteina_g": 0.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Her Calic",
    "porcion_desc": "1 cápsula",
    "kcal": 5.0,
    "proteina_g": 0.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Omeg 3",
    "porcion_desc": "1 cápsula",
    "kcal": 10.0,
    "proteina_g": 0.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Xtra Cal",
    "porcion_desc": "1 cápsula",
    "kcal": 10.0,
    "proteina_g": 0.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Queso Cheddar",
    "porcion_desc": "10 laminas (20g)",
    "kcal": 270.0,
    "proteina_g": 18.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Queso Edam",
    "porcion_desc": "100g",
    "kcal": 350.0,
    "proteina_g": 25.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Queso Mozzarella",
    "porcion_desc": "100g",
    "kcal": 280.0,
    "proteina_g": 28.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Yogurt Griego",
    "porcion_desc": "1 vaso (250ml)",
    "kcal": 136.0,
    "proteina_g": 7.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Yogurt Griego Light",
    "porcion_desc": "1 vaso (250ml)",
    "kcal": 125.0,
    "proteina_g": 15.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Yogurt Natural",
    "porcion_desc": "1 vaso (250ml)",
    "kcal": 118.0,
    "proteina_g": 6.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Leche evaporada (diluida)",
    "porcion_desc": "1 vaso (250ml)",
    "kcal": 155.0,
    "proteina_g": 8.8,
    "hidr_ml": 0
  },
  {
    "nombre": "Leche de vaca (deslactosada)",
    "porcion_desc": "1 vaso (250ml)",
    "kcal": 115.0,
    "proteina_g": 8.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Leche de soya (sin azúcar)",
    "porcion_desc": "1 vaso (250ml)",
    "kcal": 95.0,
    "proteina_g": 7.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Leche de almendras (sin azúcar)",
    "porcion_desc": "1 vaso (250ml)",
    "kcal": 35.0,
    "proteina_g": 1.2,
    "hidr_ml": 0
  },
  {
    "nombre": "Tomate (unidad)",
    "porcion_desc": "1 unidad",
    "kcal": 18.0,
    "proteina_g": 0.9,
    "hidr_ml": 0
  },
  {
    "nombre": "Manzana (unidad)",
    "porcion_desc": "1 unidad",
    "kcal": 95.0,
    "proteina_g": 0.5,
    "hidr_ml": 0
  },
  {
    "nombre": "Pera (unidad)",
    "porcion_desc": "1 unidad",
    "kcal": 101.0,
    "proteina_g": 0.6,
    "hidr_ml": 0
  },
  {
    "nombre": "Uva",
    "porcion_desc": "1 puñado (80g)",
    "kcal": 55.0,
    "proteina_g": 0.6,
    "hidr_ml": 0
  },
  {
    "nombre": "Plátano (unidad)",
    "porcion_desc": "1 unidad",
    "kcal": 105.0,
    "proteina_g": 1.3,
    "hidr_ml": 0
  },
  {
    "nombre": "Pan de molde blanco",
    "porcion_desc": "1 tajada",
    "kcal": 80.0,
    "proteina_g": 3.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Pan de molde integral",
    "porcion_desc": "1 tajada",
    "kcal": 80.0,
    "proteina_g": 3.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Pan francés",
    "porcion_desc": "1 unidad",
    "kcal": 160.0,
    "proteina_g": 5.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Pan ciabatta",
    "porcion_desc": "1 unidad",
    "kcal": 260.0,
    "proteina_g": 9.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Arándanos",
    "porcion_desc": "1 puñado (70g)",
    "kcal": 40.0,
    "proteina_g": 0.5,
    "hidr_ml": 0
  },
  {
    "nombre": "Frambuesas",
    "porcion_desc": "1 puñado (70g)",
    "kcal": 32.0,
    "proteina_g": 0.7,
    "hidr_ml": 0
  },
  {
    "nombre": "Zarzamoras",
    "porcion_desc": "1 puñado (70g)",
    "kcal": 43.0,
    "proteina_g": 1.4,
    "hidr_ml": 0
  },
  {
    "nombre": "Lechuga",
    "porcion_desc": "1 taza (50g)",
    "kcal": 8.0,
    "proteina_g": 0.5,
    "hidr_ml": 0
  },
  {
    "nombre": "Yogurt Danlac Griego Natural",
    "porcion_desc": "1 vaso (250ml)",
    "kcal": 136.0,
    "proteina_g": 7.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Yogurt Danlac Descremado Natural",
    "porcion_desc": "1 vaso (250ml)",
    "kcal": 126.0,
    "proteina_g": 5.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Chaufa de Pollo (Carta)",
    "porcion_desc": "1 plato",
    "kcal": 750.0,
    "proteina_g": 32.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Chaufa de Pollo (Menu)",
    "porcion_desc": "1 plato",
    "kcal": 420.0,
    "proteina_g": 18.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Chaufa Especial (Menu)",
    "porcion_desc": "1 plato",
    "kcal": 480.0,
    "proteina_g": 22.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Chaufa Especial (Carta)",
    "porcion_desc": "1 plato",
    "kcal": 850.0,
    "proteina_g": 40.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Arroz Aeropuerto (Menu)",
    "porcion_desc": "1 plato",
    "kcal": 600.0,
    "proteina_g": 20.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Arroz Aeropuerto (Carta)",
    "porcion_desc": "1 plato",
    "kcal": 1100.0,
    "proteina_g": 38.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Tallarín Saltado de Pollo (Menu)",
    "porcion_desc": "1 plato",
    "kcal": 400.0,
    "proteina_g": 16.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Tallarín Saltado de Pollo (Carta)",
    "porcion_desc": "1 plato",
    "kcal": 700.0,
    "proteina_g": 30.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Tallarín Taypa (Menu)",
    "porcion_desc": "1 plato",
    "kcal": 500.0,
    "proteina_g": 23.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Tallarín Taypa (Carta)",
    "porcion_desc": "1 plato",
    "kcal": 900.0,
    "proteina_g": 42.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Juane",
    "porcion_desc": "1 unidad",
    "kcal": 407.0,
    "proteina_g": 20.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Pollo Chi Jau Kay (Menu)",
    "porcion_desc": "1 plato",
    "kcal": 400.0,
    "proteina_g": 26.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Pollo Chi Jau Kay (Carta)",
    "porcion_desc": "1 plato",
    "kcal": 800.0,
    "proteina_g": 45.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Pollo TiPaKay (Menu)",
    "porcion_desc": "1 plato",
    "kcal": 420.0,
    "proteina_g": 25.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Pollo TiPaKay (Carta)",
    "porcion_desc": "1 plato",
    "kcal": 850.0,
    "proteina_g": 42.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Chancho con tamarindo (Menu)",
    "porcion_desc": "1 plato",
    "kcal": 410.0,
    "proteina_g": 23.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Chancho con tamarindo (Carta)",
    "porcion_desc": "1 plato",
    "kcal": 780.0,
    "proteina_g": 38.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Pollo Enrollado (Menu)",
    "porcion_desc": "1 plato",
    "kcal": 450.0,
    "proteina_g": 30.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Pollo Enrollado (Carta)",
    "porcion_desc": "1 plato",
    "kcal": 650.0,
    "proteina_g": 45.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Lomo en Ostión (Menu)",
    "porcion_desc": "1 plato",
    "kcal": 380.0,
    "proteina_g": 25.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Lomo en Ostión (Carta)",
    "porcion_desc": "1 plato",
    "kcal": 720.0,
    "proteina_g": 40.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Sopa Wantan (Menu)",
    "porcion_desc": "1 plato",
    "kcal": 130.0,
    "proteina_g": 6.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Sopa Wantan (Carta)",
    "porcion_desc": "1 plato",
    "kcal": 250.0,
    "proteina_g": 12.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Ceviche (Completo)",
    "porcion_desc": "1 plato",
    "kcal": 450.0,
    "proteina_g": 38.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Ceviche (Sin Guarniciones)",
    "porcion_desc": "1 plato",
    "kcal": 200.0,
    "proteina_g": 35.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Chifa Casa",
    "porcion_desc": "1 unidad",
    "kcal": 450.0,
    "proteina_g": 28.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Whopper BK",
    "porcion_desc": "1 unidad",
    "kcal": 556.0,
    "proteina_g": 22.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Whopper Cheese BK",
    "porcion_desc": "1 unidad",
    "kcal": 594.0,
    "proteina_g": 26.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Whopper Doble BK",
    "porcion_desc": "1 unidad",
    "kcal": 899.0,
    "proteina_g": 53.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Whopper Doble Cheese BK",
    "porcion_desc": "1 unidad",
    "kcal": 937.0,
    "proteina_g": 57.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Chicken Burger BK",
    "porcion_desc": "1 unidad",
    "kcal": 453.0,
    "proteina_g": 15.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Stacker BK",
    "porcion_desc": "1 unidad",
    "kcal": 462.0,
    "proteina_g": 22.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Papas Fritas BK",
    "porcion_desc": "1 unidad",
    "kcal": 328.0,
    "proteina_g": 4.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Papas Fritas Grandes BK",
    "porcion_desc": "1 unidad",
    "kcal": 430.0,
    "proteina_g": 5.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Gaseosa 21oz",
    "porcion_desc": "1 unidad",
    "kcal": 242.0,
    "proteina_g": 0.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Gaseosa 16oz",
    "porcion_desc": "1 unidad",
    "kcal": 184.0,
    "proteina_g": 0.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Gaseosa 20oz",
    "porcion_desc": "1 unidad",
    "kcal": 230.0,
    "proteina_g": 0.0,
    "hidr_ml": 0
  },
  {
    "nombre": "MC Donalds Classic",
    "porcion_desc": "1 unidad",
    "kcal": 390.0,
    "proteina_g": 12.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Mc Donalds Cheese",
    "porcion_desc": "1 unidad",
    "kcal": 320.0,
    "proteina_g": 13.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Mc Donalds Doble Cheese",
    "porcion_desc": "1 unidad",
    "kcal": 450.0,
    "proteina_g": 23.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Mc Donalds Doble con Queso",
    "porcion_desc": "1 unidad",
    "kcal": 410.0,
    "proteina_g": 22.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Mc Donalds Chicken",
    "porcion_desc": "1 unidad",
    "kcal": 470.0,
    "proteina_g": 15.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Mc Donalds Nuggets 10",
    "porcion_desc": "1 unidad",
    "kcal": 530.0,
    "proteina_g": 26.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Mc Donalds Papas",
    "porcion_desc": "1 unidad",
    "kcal": 340.0,
    "proteina_g": 3.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Pollo a la brasa (1/4 sin guarnición)",
    "porcion_desc": "1 unidad",
    "kcal": 350.0,
    "proteina_g": 20.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Pollo a la brasa 1/4 (completo)",
    "porcion_desc": "1 unidad",
    "kcal": 1500.0,
    "proteina_g": 20.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Chicharrones Caseros",
    "porcion_desc": "1 unidad",
    "kcal": 680.0,
    "proteina_g": 45.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Tocino",
    "porcion_desc": "1 unidad",
    "kcal": 45.0,
    "proteina_g": 3.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Limonada (con azucar)",
    "porcion_desc": "1 vaso (250ml)",
    "kcal": 120.0,
    "proteina_g": 0.5,
    "hidr_ml": 0
  },
  {
    "nombre": "Limonada (sin azucar)",
    "porcion_desc": "1 vaso (250ml)",
    "kcal": 30.0,
    "proteina_g": 0.5,
    "hidr_ml": 0
  },
  {
    "nombre": "Chicha (sin azucar)",
    "porcion_desc": "1 vaso (250ml)",
    "kcal": 35.0,
    "proteina_g": 0.5,
    "hidr_ml": 0
  },
  {
    "nombre": "Chicha (Con azucar)",
    "porcion_desc": "1 vaso (250ml)",
    "kcal": 100.0,
    "proteina_g": 0.5,
    "hidr_ml": 0
  },
  {
    "nombre": "Maracuya (sin azucar)",
    "porcion_desc": "1 vaso (250ml)",
    "kcal": 35.0,
    "proteina_g": 0.5,
    "hidr_ml": 0
  },
  {
    "nombre": "Maracuya (con azucar)",
    "porcion_desc": "1 vaso (250ml)",
    "kcal": 110.0,
    "proteina_g": 0.5,
    "hidr_ml": 0
  },
  {
    "nombre": "Emoliente (sin azucar)",
    "porcion_desc": "1 vaso (250ml)",
    "kcal": 15.0,
    "proteina_g": 0.5,
    "hidr_ml": 0
  },
  {
    "nombre": "Emoliente (con azucar)",
    "porcion_desc": "1 vaso (250ml)",
    "kcal": 15.0,
    "proteina_g": 0.5,
    "hidr_ml": 0
  },
  {
    "nombre": "Wantan Frito",
    "porcion_desc": "1 unidad",
    "kcal": 45.0,
    "proteina_g": 1.2,
    "hidr_ml": 0
  },
  {
    "nombre": "Siu Mai",
    "porcion_desc": "1 unidad",
    "kcal": 45.0,
    "proteina_g": 2.5,
    "hidr_ml": 0
  },
  {
    "nombre": "Palta",
    "porcion_desc": "1 unidad (200g)",
    "kcal": 200.0,
    "proteina_g": 2.5,
    "hidr_ml": 0
  },
  {
    "nombre": "Tamal",
    "porcion_desc": "1 unidad (200g)",
    "kcal": 400.0,
    "proteina_g": 10.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Arroz con pato",
    "porcion_desc": "1 plato",
    "kcal": 850.0,
    "proteina_g": 35.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Cabrito",
    "porcion_desc": "1 plato",
    "kcal": 950.0,
    "proteina_g": 45.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Arroz con pollo",
    "porcion_desc": "1 plato",
    "kcal": 750.0,
    "proteina_g": 30.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Aji de gallina",
    "porcion_desc": "1 plato",
    "kcal": 750.0,
    "proteina_g": 28.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Seco de Chabelo",
    "porcion_desc": "1 plato",
    "kcal": 1000.0,
    "proteina_g": 40.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Sudado",
    "porcion_desc": "1 plato",
    "kcal": 550.0,
    "proteina_g": 38.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Queso Fresco",
    "porcion_desc": "1 tajada (30g)",
    "kcal": 75.0,
    "proteina_g": 5.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Yuca sancochada",
    "porcion_desc": "1 unidad (100g)",
    "kcal": 120.0,
    "proteina_g": 1.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Yuca frita",
    "porcion_desc": "1 unidad (100g)",
    "kcal": 250.0,
    "proteina_g": 1.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Jamon de cerdo",
    "porcion_desc": "1 tajada (15g)",
    "kcal": 25.0,
    "proteina_g": 2.5,
    "hidr_ml": 0
  },
  {
    "nombre": "Jamon de Pavo",
    "porcion_desc": "1 tajada (15g)",
    "kcal": 15.0,
    "proteina_g": 2.5,
    "hidr_ml": 0
  },
  {
    "nombre": "Frijoles cocidos",
    "porcion_desc": "1 plato (200g)",
    "kcal": 345.0,
    "proteina_g": 16.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Anticucho",
    "porcion_desc": "1 palo (sin acompañamiento)",
    "kcal": 150.0,
    "proteina_g": 15.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Maca (Carretilla)",
    "porcion_desc": "1 vaso",
    "kcal": 215.0,
    "proteina_g": 6.5,
    "hidr_ml": 0
  },
  {
    "nombre": "Soya (Carretilla)",
    "porcion_desc": "1 vaso",
    "kcal": 150.0,
    "proteina_g": 9.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Kiwicha (Carretilla)",
    "porcion_desc": "1 vaso",
    "kcal": 185.0,
    "proteina_g": 6.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Fibra Activa",
    "porcion_desc": "1 scoop",
    "kcal": 27.0,
    "proteina_g": 0.2,
    "hidr_ml": 0
  },
  {
    "nombre": "Champiñon",
    "porcion_desc": "1 unidad",
    "kcal": 5.0,
    "proteina_g": 0.7,
    "hidr_ml": 0
  },
  {
    "nombre": "Ciruela",
    "porcion_desc": "1 unidad",
    "kcal": 23.0,
    "proteina_g": 0.3,
    "hidr_ml": 0
  },
  {
    "nombre": "Pierna de pollo",
    "porcion_desc": "1 unidad",
    "kcal": 360.0,
    "proteina_g": 34.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Granola",
    "porcion_desc": "1 taza",
    "kcal": 360.0,
    "proteina_g": 8.0,
    "hidr_ml": 0
  },
  {
    "nombre": "Granola",
    "porcion_desc": "1 cda",
    "kcal": 45.0,
    "proteina_g": 1.0,
    "hidr_ml": 0
  },
]

# =========================
# Estado inicial
# =========================
if "fecha" not in st.session_state:
    st.session_state.fecha = date.today()

if "perfil" not in st.session_state:
    st.session_state.perfil = None

if "requerimientos" not in st.session_state:
    st.session_state.requerimientos = {"kcal_obj": 0.0, "prot_obj": 0.0, "agua_obj": 0.0}

if "diario" not in st.session_state:
    st.session_state.diario = []  # {hora,nombre,porciones,porcion_desc,kcal,proteina_g,hidr_ml}

# Reset automático si cambia el día
if st.session_state.fecha != date.today():
    st.session_state.perfil = None
    st.session_state.requerimientos = {"kcal_obj": 0.0, "prot_obj": 0.0, "agua_obj": 0.0}
    st.session_state.diario = []
    st.session_state.fecha = date.today()

# =========================
# Helpers
# =========================
OBJETIVOS = [
    "Perder peso",
    "Tonificar",
    "Aumentar masa muscular",
    "Mejorar rendimiento físico",
    "Mejorar salud",
]

def es_objetivo_alto(obj: str) -> bool:
    normal = obj.strip().lower()
    return normal in {"aumentar masa muscular", "mejorar rendimiento físico", "mejorar rendimiento fisico"}

def calcular_requerimientos(peso_kg: float, altura_cm: float, edad: int, genero: str, objetivo: str):
    # Proteína
    if genero == "Hombre":
        prot = peso_kg * (2.0 if es_objetivo_alto(objetivo) else 1.6)
    else:
        prot = peso_kg * (1.8 if es_objetivo_alto(objetivo) else 1.4)

    # Calorías (Mifflin-St Jeor) +/- 250 con piso 1200 cuando aplica
    if genero == "Hombre":
        bmr = (10 * peso_kg) + (6.25 * altura_cm) - (5 * edad) + 5
    else:
        bmr = (10 * peso_kg) + (6.25 * altura_cm) - (5 * edad) - 161

    if es_objetivo_alto(objetivo):
        kcal = bmr + 250
    else:
        kcal = bmr - 250
        if kcal < 1200:
            kcal = 1200

    agua_ml = (peso_kg / 7.0) * 250.0
    return round(kcal, 0), round(prot, 1), round(agua_ml, 0)

def totales_diarios():
    kcal = sum(x["kcal"] for x in st.session_state.diario)
    prot = sum(x["proteina_g"] for x in st.session_state.diario)
    agua = sum(x["hidr_ml"] for x in st.session_state.diario)
    return kcal, prot, agua

def etiqueta_color(valor: float, objetivo: float, regla: str) -> str:
    if objetivo <= 0:
        return "⚪"
    if regla == "menor_mejor":
        return "🟢" if valor <= objetivo else "🔴"
    else:
        return "🟢" if valor >= objetivo else "🔴"

# Color para pintar números (verde si cumple, rojo si no)
def color_hex(valor: float, objetivo: float, regla: str) -> str:
    if objetivo <= 0:
        return "#ffffff"
    if regla == "menor_mejor":
        return "#2ecc71" if valor <= objetivo else "#e74c3c"
    else:
        return "#2ecc71" if valor >= objetivo else "#e74c3c"

def agregar_fila(nombre: str, porciones: float, item: Dict|None, hidr_ml_override: float | None = None):
    ahora = datetime.now().strftime(HORA_FMT)
    kcal = (item["kcal"] * porciones) if item else 0.0
    prot = (item["proteina_g"] * porciones) if item else 0.0
    hidr = (item["hidr_ml"] * porciones) if item else 0.0
    if hidr_ml_override is not None:
        hidr = float(hidr_ml_override)
    st.session_state.diario.append({
        "hora": ahora,
        "nombre": nombre,
        "porciones": porciones,
        "porcion_desc": (item["porcion_desc"] if item else f"{int(hidr)} ml"),
        "kcal": round(kcal, 1),
        "proteina_g": round(prot, 1),
        "hidr_ml": round(hidr, 0),
    })

def construir_df_diario() -> pd.DataFrame:
    if not st.session_state.diario:
        return pd.DataFrame(columns=["hora","nombre","porciones","porcion_desc","kcal","proteina_g","hidr_ml"])
    return pd.DataFrame(st.session_state.diario)

# =========================
# UI - Título
# =========================
st.markdown(
    """
    <h1 style='text-align: center; color: white; font-weight: 800; margin-top: 10px; margin-bottom: 25px;'>
        📒 Diario de Comidas
    </h1>
    """,
    unsafe_allow_html=True
)

# =========================
# 1) Perfil y requerimientos
# =========================
with st.expander("1) Datos de la persona y requerimientos", expanded=True):
    with st.form("form_perfil"):
        col1, col2, col3, col4, col5 = st.columns([1,1,1,1,1])
        with col1:
            peso = st.number_input("Peso (kg)", min_value=20.0, max_value=300.0, step=0.5, value=70.0)
        with col2:
            altura = st.number_input("Altura (cm)", min_value=120.0, max_value=220.0, step=0.5, value=170.0,
                                     help="Necesaria para el cálculo de calorías (Mifflin-St Jeor).")
        with col3:
            edad = st.number_input("Edad (años)", min_value=10, max_value=100, step=1, value=30)
        with col4:
            genero = st.selectbox("Género", ["Hombre", "Mujer"])
        with col5:
            objetivo = st.selectbox("Objetivo", OBJETIVOS)

        submit = st.form_submit_button("Calcular requerimientos")
        if submit:
            kcal_obj, prot_obj, agua_obj = calcular_requerimientos(peso, altura, edad, genero, objetivo)
            st.session_state.perfil = {
                "peso_kg": peso, "altura_cm": altura, "edad": edad, "genero": genero, "objetivo": objetivo
            }
            st.session_state.requerimientos = {
                "kcal_obj": float(kcal_obj),
                "prot_obj": float(prot_obj),
                "agua_obj": float(agua_obj),
            }

    if st.session_state.perfil:
        c1, c2, c3, c4, c5 = st.columns(5)
        with c1:
            st.metric("🎯 Calorías objetivo", f"{st.session_state.requerimientos['kcal_obj']:.0f} kcal")
        with c2:
            st.metric("🎯 Proteína objetivo", f"{st.session_state.requerimientos['prot_obj']:.1f} g")
        with c3:
            st.metric("🎯 Hidratación objetivo", f"{st.session_state.requerimientos['agua_obj']:.0f} ml")
        with c4:
            st.write(" ")
        with c5:
            if st.button("🔄 Reiniciar día", use_container_width=True, type="secondary"):
                st.session_state.diario = []
                st.session_state.fecha = date.today()

# =========================
# 2) Barras de progreso (número acumulado coloreado)
# =========================
if st.session_state.perfil:
    kcal_tot, prot_tot, agua_tot = totales_diarios()
    kcal_obj = st.session_state.requerimientos["kcal_obj"]
    prot_obj = st.session_state.requerimientos["prot_obj"]
    agua_obj = st.session_state.requerimientos["agua_obj"]

    st.subheader("Progreso del día")
    pc1, pc2, pc3 = st.columns(3)

    with pc1:
        ratio = min(1.0, (kcal_tot / kcal_obj) if kcal_obj > 0 else 0.0)
        st.progress(ratio, text=f"Calorías: {kcal_tot:.0f}/{kcal_obj:.0f} kcal ({ratio*100:.0f}%)")
        col = color_hex(kcal_tot, kcal_obj, "menor_mejor")
        st.markdown(
            f"**Acumulado:** <span style='color:{col};font-weight:700'>{kcal_tot:.0f}</span> kcal",
            unsafe_allow_html=True
        )

    with pc2:
        ratio = min(1.0, (prot_tot / prot_obj) if prot_obj > 0 else 0.0)
        st.progress(ratio, text=f"Proteína: {prot_tot:.1f}/{prot_obj:.1f} g ({ratio*100:.0f}%)")
        col = color_hex(prot_tot, prot_obj, "mayor_mejor")
        st.markdown(
            f"**Acumulado:** <span style='color:{col};font-weight:700'>{prot_tot:.1f}</span> g",
            unsafe_allow_html=True
        )

    with pc3:
        ratio = min(1.0, (agua_tot / agua_obj) if agua_obj > 0 else 0.0)
        st.progress(ratio, text=f"Hidratación: {agua_tot:.0f}/{agua_obj:.0f} ml ({ratio*100:.0f}%)")
        col = color_hex(agua_tot, agua_obj, "mayor_mejor")
        st.markdown(
            f"**Acumulado:** <span style='color:{col};font-weight:700'>{agua_tot:.0f}</span> ml",
            unsafe_allow_html=True
        )

# =========================
# 3) Registro de comidas (solo select + porciones)
# =========================
st.subheader("Registro de comidas")

def etiqueta_item(it: Dict) -> str:
    return f"{it['nombre']} — {it['porcion_desc']} · {it['kcal']} kcal · {it['proteina_g']} g prot"

lista_items = BASE_INTERNA[:]
labels = [etiqueta_item(it) for it in lista_items]

with st.form("form_comida", clear_on_submit=True):
    col_a, col_b, col_c = st.columns([3,1,1])
    with col_a:
        sel_label = st.selectbox("Elige un alimento/bebida", labels)
        idx = labels.index(sel_label)
        item = lista_items[idx]
        st.caption(f"Porción base: **{item['porcion_desc']}**")
    with col_b:
        porciones = st.number_input("Cant. porciones", min_value=0.25, max_value=20.0, step=0.25, value=1.0)
    with col_c:
        st.write(" ")
        add_food = st.form_submit_button("➕ Agregar", use_container_width=True)

    if add_food and item:
        agregar_fila(nombre=item["nombre"], porciones=porciones, item=item)
        st.rerun()  # <-- fuerza actualización inmediata

# Registro rápido de agua
with st.form("form_agua", clear_on_submit=True):
    colx, coly = st.columns([2,1])
    with colx:
        agua_ml_in = st.number_input("Registrar agua (ml)", min_value=0, max_value=5000, step=50, value=0)
    with coly:
        st.write(" ")
        add_agua = st.form_submit_button("💧 Agregar agua", use_container_width=True)
    if add_agua and agua_ml_in > 0:
        agregar_fila(nombre="Agua (ml)", porciones=1.0, item=None, hidr_ml_override=agua_ml_in)
        st.rerun()  # <-- fuerza actualización inmediata

# =========================
# 4) Tabla del día y acciones (selección de filas)
# =========================
df = construir_df_diario()

sel_rows = []
if not df.empty:
    # ---- NUEVO: agregar ratio kcal/g proteína (None si prot <= 0) ----
    df = df.copy()
    df["kcal_por_g_prot"] = df.apply(
        lambda r: round(r["kcal"] / r["proteina_g"], 2) if r["proteina_g"] and r["proteina_g"] > 0 else None,
        axis=1
    )
    # ------------------------------------------------------------------

    df_display = df.copy()
    df_display.insert(0, "Seleccionar", False)

    edited = st.data_editor(
        df_display,
        use_container_width=True,
        hide_index=True,
        disabled={
            "hora": True,
            "nombre": True,
            "porciones": True,
            "porcion_desc": True,
            "kcal": True,
            "proteina_g": True,
            "hidr_ml": True,
            "kcal_por_g_prot": True,  # <- NUEVO: columna calculada bloqueada
        },
        key="editor_diario"
    )
    sel_rows = edited.index[edited["Seleccionar"] == True].tolist()
else:
    st.dataframe(df, use_container_width=True, hide_index=True)

col_btn1, col_btn2, _ = st.columns([1,1,2])
with col_btn1:
    if st.button("🗑️ Eliminar fila seleccionada", use_container_width=True, type="secondary", disabled=(len(sel_rows)==0)):
        for i in sorted(sel_rows, reverse=True):
            if 0 <= i < len(st.session_state.diario):
                st.session_state.diario.pop(i)
        st.rerun()

with col_btn2:
    if st.button("🗑️ Vaciar diario", use_container_width=True, type="secondary", disabled=df.empty):
        st.session_state.diario = []
        st.rerun()

# =========================
# 5) Resumen + Exportar (número acumulado coloreado)
# =========================
if st.session_state.perfil:
    kcal_tot, prot_tot, agua_tot = totales_diarios()
    kcal_obj = st.session_state.requerimientos["kcal_obj"]
    prot_obj = st.session_state.requerimientos["prot_obj"]
    agua_obj = st.session_state.requerimientos["agua_obj"]

    st.markdown("---")
    c1, c2, c3 = st.columns(3)
    with c1:
        col = color_hex(kcal_tot, kcal_obj, "menor_mejor")
        st.markdown(
            f"### Calorías\n<span style='font-weight:700;color:{col}'>{kcal_tot:.0f}</span> / {kcal_obj:.0f} kcal",
            unsafe_allow_html=True
        )
    with c2:
        col = color_hex(prot_tot, prot_obj, "mayor_mejor")
        st.markdown(
            f"### Proteína\n<span style='font-weight:700;color:{col}'>{prot_tot:.1f}</span> / {prot_obj:.1f} g",
            unsafe_allow_html=True
        )
    with c3:
        col = color_hex(agua_tot, agua_obj, "mayor_mejor")
        st.markdown(
            f"### Hidratación\n<span style='font-weight:700;color:{col}'>{agua_tot:.0f}</span> / {agua_obj:.0f} ml",
            unsafe_allow_html=True
        )

    # Datos para exportar
    df_export = construir_df_diario()
    resumen = pd.DataFrame({
        "Métrica": ["Calorías (kcal)", "Proteína (g)", "Hidratación (ml)"],
        "Total del día": [round(kcal_tot,0), round(prot_tot,1), round(agua_tot,0)],
        "Objetivo": [round(kcal_obj,0), round(prot_obj,1), round(agua_obj,0)],
    })
    perfil = pd.DataFrame([{
        "Peso (kg)": st.session_state.perfil["peso_kg"],
        "Altura (cm)": st.session_state.perfil["altura_cm"],
        "Edad": st.session_state.perfil["edad"],
        "Género": st.session_state.perfil["genero"],
        "Objetivo": st.session_state.perfil["objetivo"],
        "Fecha": st.session_state.fecha.strftime("%Y-%m-%d"),
    }])

    # Detectar engine disponible para Excel
    excel_engine = None
    try:
        import xlsxwriter  # noqa: F401
        excel_engine = "xlsxwriter"
    except Exception:
        try:
            import openpyxl  # noqa: F401
            excel_engine = "openpyxl"
        except Exception:
            excel_engine = None

    if excel_engine:
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine=excel_engine) as writer:
            df_export.to_excel(writer, index=False, sheet_name="Diario")
            resumen.to_excel(writer, index=False, sheet_name="Resumen")

            ws = writer.sheets["Resumen"]
            start_row = len(resumen) + 2

            if excel_engine == "xlsxwriter":
                for col_idx, col_name in enumerate(perfil.columns):
                    ws.write(start_row, col_idx, col_name)
                for row_idx in range(len(perfil)):
                    for col_idx, col_name in enumerate(perfil.columns):
                        ws.write(start_row + 1 + row_idx, col_idx, perfil.iloc[row_idx][col_name])
            else:
                from openpyxl.utils import get_column_letter  # noqa: F401
                for j, col_name in enumerate(perfil.columns, start=1):
                    ws.cell(row=start_row+1, column=j, value=col_name)
                for r in range(len(perfil)):
                    for j, col_name in enumerate(perfil.columns, start=1):
                        ws.cell(row=start_row+2+r, column=j, value=perfil.iloc[r][col_name])

        st.download_button(
            label="📥 Exportar Diario (Excel)",
            data=buffer.getvalue(),
            file_name=f"diario_comidas_{st.session_state.fecha.isoformat()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    else:
        # Fallback: ZIP con CSVs si no hay engines de Excel
        import zipfile
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
            zf.writestr("Diario.csv", df_export.to_csv(index=False))
            zf.writestr("Resumen.csv", resumen.to_csv(index=False))
            zf.writestr("Perfil.csv", perfil.to_csv(index=False))
        st.warning("No se encontró `xlsxwriter` ni `openpyxl`. Se exportará un ZIP con CSVs.")
        st.download_button(
            label="📥 Exportar Diario (ZIP con CSVs)",
            data=zip_buffer.getvalue(),
            file_name=f"diario_comidas_{st.session_state.fecha.isoformat()}.zip",
            mime="application/zip",
            use_container_width=True
        )
else:
    st.info("Primero completa tus datos y calcula tus requerimientos para activar el seguimiento y la exportación.")
